"""Validated spreadsheet configuration for Pricing Dashboard security roots."""

from __future__ import annotations

import csv
from datetime import date, datetime
import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

SHEET_NAME = "Security Roots"
UPDATE_SHEET_NAME = "Bloomberg Update"
ROOT_COLUMNS = (
    "enabled",
    "root",
    "common_name",
    "yellow_key",
    "native_unit",
    "bbl_per_mt",
    "gal_per_bbl",
    "ticker_template",
    "tradingview_symbol",
    "aliases",
    "product_group",
    "sort_order",
)
VALID_YELLOW_KEYS = {"comdty": "Comdty", "index": "Index"}
VALID_UNITS = ("cpg", "$/gal", "$/bbl", "$/MT")
UNIT_ALIASES = {
    "cpg": "cpg",
    "centpergallon": "cpg",
    "centspergallon": "cpg",
    "cents/gal": "cpg",
    "$/gal": "$/gal",
    "usd/gal": "$/gal",
    "dollarspergallon": "$/gal",
    "dollarpergallon": "$/gal",
    "$/bbl": "$/bbl",
    "usd/bbl": "$/bbl",
    "dollarsperbarrel": "$/bbl",
    "dollarperbarrel": "$/bbl",
    "$/mt": "$/MT",
    "usd/mt": "$/MT",
    "dollarspermetricton": "$/MT",
    "dollarpermetricton": "$/MT",
}
TRUE_VALUES = {"1", "true", "yes", "y", "on", "enabled"}
FALSE_VALUES = {"0", "false", "no", "n", "off", "disabled"}
ROOT_PATTERN = re.compile(r"^[A-Z0-9][A-Z0-9._-]*$")
FIELD_PATTERN = re.compile(r"^[A-Z][A-Z0-9_]*$")


class ConfigValidationError(ValueError):
    """Raised with all actionable spreadsheet problems at once."""

    def __init__(self, issues: Iterable[str]):
        self.issues = tuple(str(issue) for issue in issues)
        super().__init__("Security root configuration is invalid:\n- " + "\n- ".join(self.issues))


@dataclass(frozen=True)
class SecurityRoot:
    enabled: bool
    root: str
    display_name: str
    yellow_key: str
    native_unit: str
    bbl_per_mt: float
    gal_per_bbl: float
    ticker_template: str
    tradingview_symbol: str
    aliases: tuple[str, ...]
    product_group: str
    sort_order: int

    @property
    def common_name(self) -> str:
        """Authoritative dashboard label configured in the spreadsheet."""

        return self.display_name

    def to_dict(self) -> dict[str, Any]:
        return {
            "enabled": self.enabled,
            "root": self.root,
            "common_name": self.common_name,
            # Retained for the existing browser payload contract.
            "display_name": self.display_name,
            "yellow_key": self.yellow_key,
            "native_unit": self.native_unit,
            "bbl_per_mt": self.bbl_per_mt,
            "gal_per_bbl": self.gal_per_bbl,
            "ticker_template": self.ticker_template,
            "tradingview_symbol": self.tradingview_symbol,
            "aliases": list(self.aliases),
            "product_group": self.product_group,
            "sort_order": self.sort_order,
        }


@dataclass(frozen=True)
class UpdateSettings:
    """Owner-only Bloomberg update settings stored beside the security roots."""

    history_start: date
    contract_start_year: int
    contract_end_year: int
    contract_history_months: int
    reference_depth: int
    overlap_days: int
    fields: tuple[str, ...]
    host: str
    port: int
    service: str
    batch_size: int
    request_timeout_seconds: int
    standalone_max_mb: float

    def to_dict(self) -> dict[str, Any]:
        return {
            "history_start": self.history_start.isoformat(),
            "contract_start_year": self.contract_start_year,
            "contract_end_year": self.contract_end_year,
            "contract_history_months": self.contract_history_months,
            "reference_depth": self.reference_depth,
            "overlap_days": self.overlap_days,
            "fields": list(self.fields),
            "host": self.host,
            "port": self.port,
            "service": self.service,
            "batch_size": self.batch_size,
            "request_timeout_seconds": self.request_timeout_seconds,
            "standalone_max_mb": self.standalone_max_mb,
        }


def default_update_settings() -> UpdateSettings:
    current_year = date.today().year
    return UpdateSettings(
        history_start=date(current_year - 7, 1, 1),
        contract_start_year=current_year - 6,
        contract_end_year=current_year + 2,
        contract_history_months=24,
        reference_depth=2,
        overlap_days=7,
        fields=("PX_LAST", "PX_CLOSE", "PX_SETTLE", "PX_FAIR_1430"),
        host="localhost",
        port=8194,
        service="//blp/refdata",
        batch_size=25,
        request_timeout_seconds=120,
        standalone_max_mb=20.0,
    )


@dataclass(frozen=True)
class RootConfig:
    roots: tuple[SecurityRoot, ...]
    source_path: Path
    update: UpdateSettings = field(default_factory=default_update_settings)

    @property
    def enabled_roots(self) -> tuple[SecurityRoot, ...]:
        return tuple(root for root in self.roots if root.enabled)

    @property
    def enabled_root_codes(self) -> tuple[str, ...]:
        return tuple(root.root for root in self.enabled_roots)

    @property
    def by_root(self) -> dict[str, SecurityRoot]:
        return {root.root: root for root in self.roots}

    @property
    def alias_to_root(self) -> dict[str, str]:
        result: dict[str, str] = {}
        for item in self.enabled_roots:
            result[item.root.upper()] = item.root
            for alias in item.aliases:
                result[alias.upper()] = item.root
        return result

    def resolve_root(self, value: object) -> str | None:
        key = str(value or "").strip().upper()
        return self.alias_to_root.get(key)

    def to_dict(self) -> dict[str, dict[str, Any]]:
        return {
            item.root: item.to_dict()
            for item in sorted(self.enabled_roots, key=lambda root: (root.sort_order, root.root))
        }


def normalize_unit(value: object) -> str:
    text = str(value or "").strip().lower().replace(" ", "").replace("_", "")
    return UNIT_ALIASES.get(text, "")


def _parse_enabled(value: object, row_number: int, issues: list[str]) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value if value is not None else "").strip().lower()
    if not text:
        return True
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    issues.append(f"row {row_number}: enabled must be TRUE or FALSE, got {value!r}")
    return False


def _positive_float(value: object, field: str, row_number: int, issues: list[str]) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        issues.append(f"row {row_number}: {field} must be a positive number")
        return 0.0
    if not math.isfinite(number) or number <= 0:
        issues.append(f"row {row_number}: {field} must be a positive finite number")
        return 0.0
    return round(number, 5)


def _sort_order(value: object, row_number: int, issues: list[str]) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        issues.append(f"row {row_number}: sort_order must be an integer")
        return row_number
    if number < 0:
        issues.append(f"row {row_number}: sort_order must be zero or greater")
    return number


def _split_aliases(value: object) -> tuple[str, ...]:
    text = str(value or "").strip()
    if not text:
        return ()
    parts = re.split(r"[|,;]", text)
    return tuple(dict.fromkeys(part.strip().upper() for part in parts if part.strip()))


def _read_csv_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = tuple(reader.fieldnames or ())
        missing = _missing_root_columns(headers)
        if missing:
            raise ConfigValidationError([f"missing spreadsheet columns: {', '.join(missing)}"])
        return [(index, dict(row)) for index, row in enumerate(reader, start=2)]


def _read_xlsx_rows(path: Path) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ConfigValidationError([f"workbook must contain a '{SHEET_NAME}' sheet"])
        sheet = workbook[SHEET_NAME]
        values = sheet.iter_rows(values_only=True)
        headers = tuple(str(value or "").strip() for value in next(values, ()))
        missing = _missing_root_columns(headers)
        if missing:
            raise ConfigValidationError([f"missing spreadsheet columns: {', '.join(missing)}"])
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values_row in enumerate(values, start=2):
            row = dict(zip(headers, values_row))
            if any(value not in (None, "") for value in row.values()):
                rows.append((row_number, row))
        return rows
    finally:
        workbook.close()


def _missing_root_columns(headers: Iterable[str]) -> list[str]:
    """Require the current schema while accepting legacy display_name files."""

    available = set(headers)
    missing = [column for column in ROOT_COLUMNS if column not in available]
    if "common_name" in missing and "display_name" in available:
        missing.remove("common_name")
    return missing


def _parse_date_setting(value: object, key: str, issues: list[str], fallback: date) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return fallback
    try:
        return date.fromisoformat(text)
    except ValueError:
        issues.append(f"{UPDATE_SHEET_NAME}: {key} must use YYYY-MM-DD")
        return fallback


def _parse_int_setting(
    values: dict[str, object],
    key: str,
    fallback: int,
    minimum: int,
    maximum: int,
    issues: list[str],
) -> int:
    raw = values.get(key, fallback)
    try:
        parsed = int(raw)
    except (TypeError, ValueError):
        issues.append(f"{UPDATE_SHEET_NAME}: {key} must be an integer")
        return fallback
    if parsed < minimum or parsed > maximum:
        issues.append(
            f"{UPDATE_SHEET_NAME}: {key} must be between {minimum} and {maximum}"
        )
        return fallback
    return parsed


def _read_update_settings(path: Path) -> UpdateSettings:
    defaults = default_update_settings()
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if UPDATE_SHEET_NAME not in workbook.sheetnames:
            return defaults
        sheet = workbook[UPDATE_SHEET_NAME]
        values: dict[str, object] = {}
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_number == 1:
                continue
            key = str(row[0] or "").strip().lower() if row else ""
            if key:
                values[key] = row[1] if len(row) > 1 else None
    finally:
        workbook.close()

    issues: list[str] = []
    history_start = _parse_date_setting(
        values.get("history_start"), "history_start", issues, defaults.history_start
    )
    contract_start_year = _parse_int_setting(
        values, "contract_start_year", defaults.contract_start_year, 1970, 2200, issues
    )
    contract_end_year = _parse_int_setting(
        values, "contract_end_year", defaults.contract_end_year, 1970, 2200, issues
    )
    contract_history_months = _parse_int_setting(
        values, "contract_history_months", defaults.contract_history_months, 1, 120, issues
    )
    reference_depth = _parse_int_setting(
        values, "reference_depth", defaults.reference_depth, 1, 10, issues
    )
    overlap_days = _parse_int_setting(
        values, "overlap_days", defaults.overlap_days, 0, 60, issues
    )
    port = _parse_int_setting(values, "port", defaults.port, 1, 65535, issues)
    batch_size = _parse_int_setting(
        values, "batch_size", defaults.batch_size, 1, 100, issues
    )
    request_timeout_seconds = _parse_int_setting(
        values,
        "request_timeout_seconds",
        defaults.request_timeout_seconds,
        5,
        3600,
        issues,
    )

    fields_raw = values.get("fields", ",".join(defaults.fields))
    fields = tuple(
        dict.fromkeys(
            token.strip().upper()
            for token in re.split(r"[,;|]", str(fields_raw or ""))
            if token.strip()
        )
    )
    if not fields:
        issues.append(f"{UPDATE_SHEET_NAME}: fields must include at least PX_LAST")
    elif "PX_LAST" not in fields:
        issues.append(f"{UPDATE_SHEET_NAME}: fields must include PX_LAST")
    invalid_fields = [item for item in fields if not FIELD_PATTERN.fullmatch(item)]
    if invalid_fields:
        issues.append(
            f"{UPDATE_SHEET_NAME}: unsupported field names: {', '.join(invalid_fields)}"
        )

    host = str(values.get("host", defaults.host) or "").strip()
    service = str(values.get("service", defaults.service) or "").strip()
    if not host:
        issues.append(f"{UPDATE_SHEET_NAME}: host is required")
    if not service.startswith("//"):
        issues.append(f"{UPDATE_SHEET_NAME}: service must start with //")
    try:
        standalone_max_mb = float(values.get("standalone_max_mb", defaults.standalone_max_mb))
    except (TypeError, ValueError):
        issues.append(f"{UPDATE_SHEET_NAME}: standalone_max_mb must be a number")
        standalone_max_mb = defaults.standalone_max_mb
    if not math.isfinite(standalone_max_mb) or standalone_max_mb <= 0:
        issues.append(f"{UPDATE_SHEET_NAME}: standalone_max_mb must be positive")

    if contract_end_year < contract_start_year:
        issues.append(
            f"{UPDATE_SHEET_NAME}: contract_end_year must be greater than or equal to contract_start_year"
        )
    if issues:
        raise ConfigValidationError(issues)
    return UpdateSettings(
        history_start=history_start,
        contract_start_year=contract_start_year,
        contract_end_year=contract_end_year,
        contract_history_months=contract_history_months,
        reference_depth=reference_depth,
        overlap_days=overlap_days,
        fields=fields,
        host=host,
        port=port,
        service=service,
        batch_size=batch_size,
        request_timeout_seconds=request_timeout_seconds,
        standalone_max_mb=round(standalone_max_mb, 5),
    )


def load_root_config(path: str | Path) -> RootConfig:
    source = Path(path).expanduser().resolve()
    if not source.exists():
        raise ConfigValidationError([f"configuration file not found: {source}"])
    suffix = source.suffix.lower()
    if suffix == ".csv":
        raw_rows = _read_csv_rows(source)
        update_settings = default_update_settings()
    elif suffix in {".xlsx", ".xlsm"}:
        raw_rows = _read_xlsx_rows(source)
        update_settings = _read_update_settings(source)
    else:
        raise ConfigValidationError(["configuration must be an .xlsx, .xlsm, or .csv file"])

    issues: list[str] = []
    roots: list[SecurityRoot] = []
    seen_roots: dict[str, int] = {}
    alias_owner: dict[str, str] = {}

    for row_number, row in raw_rows:
        enabled = _parse_enabled(row.get("enabled"), row_number, issues)
        root = str(row.get("root") or "").strip().upper()
        common_name = str(
            row.get("common_name") or row.get("display_name") or ""
        ).strip()
        yellow_raw = str(row.get("yellow_key") or "").strip().lower()
        yellow_key = VALID_YELLOW_KEYS.get(yellow_raw, "")
        native_unit = normalize_unit(row.get("native_unit"))
        bbl_per_mt = _positive_float(row.get("bbl_per_mt"), "bbl_per_mt", row_number, issues)
        gal_per_bbl = _positive_float(row.get("gal_per_bbl"), "gal_per_bbl", row_number, issues)
        ticker_template = str(row.get("ticker_template") or "").strip()
        aliases = _split_aliases(row.get("aliases"))
        product_group = str(row.get("product_group") or "Other").strip() or "Other"

        if not root:
            issues.append(f"row {row_number}: root is required")
        elif not ROOT_PATTERN.fullmatch(root):
            issues.append(f"row {row_number}: root {root!r} contains unsupported characters")
        elif root in seen_roots:
            issues.append(f"row {row_number}: duplicate root {root!r} (first seen on row {seen_roots[root]})")
        else:
            seen_roots[root] = row_number
        if not common_name:
            issues.append(f"row {row_number}: common_name is required")
        if not yellow_key:
            issues.append(f"row {row_number}: yellow_key must be Comdty or Index")
        if not native_unit:
            issues.append(
                f"row {row_number}: native_unit must be one of {', '.join(VALID_UNITS)}"
            )
        required_tokens = ("{root}", "{month_code}", "{yellow_key}")
        missing_tokens = [token for token in required_tokens if token not in ticker_template]
        year_tokens = (
            "{y}", "{year_1d}", "{year_digit}",
            "{yy}", "{year_2d}", "{year}", "{yyyy}",
        )
        if not any(token in ticker_template for token in year_tokens):
            missing_tokens.append("{y} (or {yy}/{year})")
        if missing_tokens:
            issues.append(
                f"row {row_number}: ticker_template is missing {', '.join(missing_tokens)}"
            )
        for alias in aliases:
            if alias == root:
                continue
            previous = alias_owner.get(alias)
            if previous and previous != root:
                issues.append(f"row {row_number}: alias {alias!r} is already assigned to {previous}")
            if alias in seen_roots and alias != root:
                issues.append(f"row {row_number}: alias {alias!r} collides with a root")
            alias_owner[alias] = root

        roots.append(
            SecurityRoot(
                enabled=enabled,
                root=root,
                display_name=common_name,
                yellow_key=yellow_key,
                native_unit=native_unit,
                bbl_per_mt=bbl_per_mt,
                gal_per_bbl=gal_per_bbl,
                ticker_template=ticker_template,
                tradingview_symbol=str(row.get("tradingview_symbol") or "").strip(),
                aliases=aliases,
                product_group=product_group,
                sort_order=_sort_order(row.get("sort_order"), row_number, issues),
            )
        )

    if not roots:
        issues.append("configuration has no security-root rows")
    if roots and not any(root.enabled for root in roots):
        issues.append("configuration has no enabled security roots")
    if issues:
        raise ConfigValidationError(issues)
    return RootConfig(tuple(roots), source, update_settings)
