from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.root_config import (
    ConfigValidationError,
    ROOT_COLUMNS,
    UPDATE_SHEET_NAME,
    load_root_config,
)


class RootConfigTests(unittest.TestCase):
    def test_default_workbook_contains_requested_roots_and_native_units(self) -> None:
        config = load_root_config("config/security_roots.xlsx")
        metadata = config.to_dict()
        self.assertEqual(metadata["WU"]["common_name"], "GC Jet")
        self.assertEqual(metadata["WU"]["display_name"], "GC Jet")
        self.assertEqual(metadata["WU"]["native_unit"], "cpg")
        self.assertEqual(metadata["HO"]["native_unit"], "cpg")
        self.assertEqual(metadata["QS"]["native_unit"], "$/MT")
        self.assertEqual(metadata["RVO"]["curve_mode"], "flat")
        self.assertEqual(metadata["RVO"]["ticker_template"], "{root} {yellow_key}")
        self.assertEqual(metadata["RVO"]["yellow_key"], "Index")
        self.assertEqual(config.resolve_root("ME"), "WU")
        self.assertEqual(config.update.host, "localhost")
        self.assertEqual(config.update.port, 8194)
        self.assertEqual(config.update.contract_start_year, 2020)
        self.assertEqual(config.update.contract_end_year, 2028)
        self.assertIn("PX_LAST", config.update.fields)
        self.assertEqual(config.update.dashboard_fields, ("PX_LAST",))

    def test_workbook_is_easy_to_fill_without_free_text_units(self) -> None:
        workbook = load_workbook("config/security_roots.xlsx")
        try:
            self.assertIn("Instructions", workbook.sheetnames)
            self.assertIn(UPDATE_SHEET_NAME, workbook.sheetnames)
            sheet = workbook["Security Roots"]
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertIn("common_name", [cell.value for cell in sheet[1]])
            validations = list(sheet.data_validations.dataValidation)
            self.assertGreaterEqual(len(validations), 4)
            formulas = {validation.formula1 for validation in validations}
            self.assertIn('"Comdty,Index"', formulas)
            self.assertIn('"cpg,$/gal,$/bbl,$/MT"', formulas)
            self.assertIn('"Monthly,Flat"', formulas)
            update_sheet = workbook[UPDATE_SHEET_NAME]
            update_settings = {
                str(row[0].value): row[1].value
                for row in update_sheet.iter_rows(min_row=2)
                if row[0].value
            }
            self.assertEqual(update_settings["dashboard_fields"], "PX_LAST")
        finally:
            workbook.close()

    def test_dashboard_fields_must_be_requested_bloomberg_fields(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "invalid-dashboard-fields.xlsx"
            workbook = load_workbook("config/security_roots.xlsx")
            try:
                update_sheet = workbook[UPDATE_SHEET_NAME]
                for row in update_sheet.iter_rows(min_row=2):
                    if row[0].value == "fields":
                        row[1].value = "PX_LAST"
                    elif row[0].value == "dashboard_fields":
                        row[1].value = "PX_LAST,PX_SETTLE"
                workbook.save(path)
            finally:
                workbook.close()
            with self.assertRaisesRegex(
                ConfigValidationError,
                "dashboard_fields must also appear in fields: PX_SETTLE",
            ):
                load_root_config(path)

    def test_duplicate_roots_and_invalid_units_are_reported_together(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "roots.csv"
            base = [True, "WU", "GC Jet", "Comdty", "cpg", 7.45, 42, "{root}{month_code}{yy} {yellow_key}", "Monthly", "", "", "Refined", 1]
            duplicate = [True, "wu", "Duplicate", "Index", "points", 7.45, 42, "{root}{month_code}{yy} {yellow_key}", "Monthly", "", "", "Refined", 2]
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(ROOT_COLUMNS)
                writer.writerows([base, duplicate])
            with self.assertRaises(ConfigValidationError) as raised:
                load_root_config(path)
            message = str(raised.exception)
            self.assertIn("duplicate root", message)
            self.assertIn("native_unit", message)

    def test_legacy_display_name_header_remains_supported(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "legacy_roots.csv"
            headers = [
                "display_name" if column == "common_name" else column
                for column in ROOT_COLUMNS
                if column != "curve_mode"
            ]
            row = [
                True, "WU", "Legacy Jet Name", "Comdty", "cpg", 7.45, 42,
                "{root}{month_code}{yy} {yellow_key}", "", "", "Refined", 1,
            ]
            with path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(headers)
                writer.writerow(row)

            config = load_root_config(path)
            self.assertEqual(config.enabled_roots[0].common_name, "Legacy Jet Name")
            self.assertEqual(config.enabled_roots[0].curve_mode, "monthly")


if __name__ == "__main__":
    unittest.main()
