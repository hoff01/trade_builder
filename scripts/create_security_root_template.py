#!/usr/bin/env python3
"""Create the editable security-root workbook and a review-friendly CSV mirror."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from app.root_config import (
    ROOT_COLUMNS,
    SHEET_NAME,
    UPDATE_SHEET_NAME,
    default_update_settings,
)

DEFAULT_ROWS = (
    (True, "WU", "GC Jet", "Comdty", "cpg", 7.45, 42, "{root}{month_code}{y} {yellow_key}", "Monthly", "NYMEX:WU", "ME|GC JET", "Refined Products", 10),
    (True, "HO", "Heating Oil", "Comdty", "cpg", 7.45, 42, "{root}{month_code}{y} {yellow_key}", "Monthly", "NYMEX:HO", "ULSD|HEATING OIL", "Refined Products", 20),
    (True, "RB", "RBOB Gasoline", "Comdty", "cpg", 8.33, 42, "{root}{month_code}{y} {yellow_key}", "Monthly", "NYMEX:RB", "RBOB", "Refined Products", 30),
    (True, "RVO", "RVO", "Index", "cpg", 7.45, 42, "{root} {yellow_key}", "Flat", "", "RENEWABLE VOLUME OBLIGATION", "Renewable Fuels", 35),
    (True, "QS", "ICE Low Sulphur Gasoil", "Comdty", "$/MT", 7.45, 42, "{root}{month_code}{y} {yellow_key}", "Monthly", "ICEEUR:QS1!", "GASOIL|LSGO", "Refined Products", 40),
    (True, "CL", "WTI Crude Oil", "Comdty", "$/bbl", 7.33, 42, "{root}{month_code}{y} {yellow_key}", "Monthly", "NYMEX:CL", "WTI", "Crude", 50),
    (True, "CO", "Brent Crude Oil", "Comdty", "$/bbl", 7.33, 42, "{root}{month_code}{y} {yellow_key}", "Monthly", "TVC:UKOIL", "BRENT", "Crude", 60),
)


def build_workbook(path: Path) -> None:
    update_defaults = default_update_settings()
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    instructions.append(["Pricing Dashboard — Security Root Setup"])
    instructions.append(["1", "Add one row per Bloomberg root and set common_name to the label shown everywhere in the dashboard."])
    instructions.append(["2", "Choose Comdty or Index, the native quote unit, and Monthly or Flat curve mode from the dropdowns."])
    instructions.append(["3", "Keep bbl_per_mt and gal_per_bbl explicit so every output-unit conversion is deterministic."])
    instructions.append(["4", "Review the Bloomberg Update sheet to control dates, full data fields, lightweight dashboard fields, and connection settings."])
    instructions.append(["5", "Save the workbook, start the local dashboard, and press UPDATE DATA."])
    instructions.append(["Ticker example", "HO + Feb (G) + {y} + Comdty produces HOG6 Comdty for 2026."])
    instructions.append(["Flat example", "RVO uses Flat with {root} {yellow_key}; Bloomberg pulls RVO Index once and the dashboard applies each daily value across the full curve."])
    instructions.append(["Year placeholders", "Use {y} for 6, {yy} for 26, or {year} for 2026. yellow_key independently controls Comdty vs Index."])
    instructions.append([])
    instructions.append(["Conversion checks"])
    instructions.append(["cpg → $/gal", "divide by 100"])
    instructions.append(["$/gal → $/bbl", "multiply by gal_per_bbl"])
    instructions.append(["$/bbl → $/MT", "multiply by bbl_per_mt"])
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 105
    instructions["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="1D4ED8")
    instructions.merge_cells("A1:B1")

    sheet = workbook.create_sheet(SHEET_NAME)
    sheet.append(ROOT_COLUMNS)
    for row in DEFAULT_ROWS:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:M{sheet.max_row}"
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1E3A8A")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    sheet["H1"].comment = Comment(
        "Examples: {root}{month_code}{y} {yellow_key} -> HOG6 Comdty; "
        "use {yy} for HOG26 or {year} for HOG2026. Set yellow_key separately.",
        "Pricing Dashboard",
    )

    widths = {
        "A": 11, "B": 10, "C": 28, "D": 13, "E": 14, "F": 13,
        "G": 13, "H": 43, "I": 14, "J": 24, "K": 30, "L": 22, "M": 12,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    table = Table(displayName="SecurityRoots", ref=f"A1:M{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    validations = (
        ("A2:A501", '"TRUE,FALSE"', "Choose TRUE or FALSE"),
        ("D2:D501", '"Comdty,Index"', "Choose Comdty or Index"),
        ("E2:E501", '"cpg,$/gal,$/bbl,$/MT"', "Choose the Bloomberg native quote unit"),
        ("I2:I501", '"Monthly,Flat"', "Monthly uses dated contracts; Flat uses one daily series with no month selector"),
    )
    for cell_range, formula, prompt in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.errorTitle = "Invalid value"
        validation.error = prompt
        validation.promptTitle = "Pricing Dashboard"
        validation.prompt = prompt
        validation.showErrorMessage = True
        validation.showInputMessage = True
        sheet.add_data_validation(validation)
        validation.add(cell_range)

    update_sheet = workbook.create_sheet(UPDATE_SHEET_NAME)
    update_sheet.append(("setting", "value", "what it controls"))
    update_rows = (
        ("history_start", update_defaults.history_start.isoformat(), "Earliest observation date retained in the update."),
        ("contract_start_year", update_defaults.contract_start_year, "First delivery year included in the ticker universe."),
        ("contract_end_year", update_defaults.contract_end_year, "Last delivery year included in the ticker universe."),
        ("contract_history_months", update_defaults.contract_history_months, "Months of history requested before each delivery month."),
        ("reference_depth", update_defaults.reference_depth, "Maximum dated-contract reference retained, normally 1 or 2."),
        ("overlap_days", update_defaults.overlap_days, "Days re-pulled for active contracts during an incremental update."),
        ("fields", ",".join(update_defaults.fields), "Bloomberg fields retained in the full CSV and Parquet; PX_LAST is required."),
        ("dashboard_fields", ",".join(update_defaults.dashboard_fields), "Subset embedded in the portable dashboard; every value must also appear in fields."),
        ("host", update_defaults.host, "Bloomberg Desktop API host."),
        ("port", update_defaults.port, "Bloomberg Desktop API port."),
        ("service", update_defaults.service, "Bloomberg reference-data service."),
        ("batch_size", update_defaults.batch_size, "Maximum securities sent in one Bloomberg request."),
        ("request_timeout_seconds", update_defaults.request_timeout_seconds, "Hard limit for each Bloomberg request."),
        ("standalone_max_mb", update_defaults.standalone_max_mb, "Maximum portable HTML size before publication stops."),
    )
    for row in update_rows:
        update_sheet.append(row)
    update_sheet.freeze_panes = "A2"
    update_sheet.sheet_view.showGridLines = False
    update_sheet.column_dimensions["A"].width = 30
    update_sheet.column_dimensions["B"].width = 58
    update_sheet.column_dimensions["C"].width = 82
    for cell in update_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    update_table = Table(displayName="BloombergUpdateSettings", ref=f"A1:C{update_sheet.max_row}")
    update_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    update_sheet.add_table(update_table)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_csv(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\n")
        writer.writerow(ROOT_COLUMNS)
        writer.writerows(DEFAULT_ROWS)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", default="config/security_roots.xlsx")
    parser.add_argument("--csv", default="config/security_roots.example.csv")
    args = parser.parse_args()
    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.csv)
    build_workbook(xlsx_path)
    write_csv(csv_path)
    print(f"wrote={xlsx_path}")
    print(f"wrote={csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
